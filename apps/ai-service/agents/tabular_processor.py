"""Tabular Processing Agent - parses CSV, XLSX, Parquet, and other tabular formats."""

from __future__ import annotations

import io
import logging
import os
from typing import Any

from core.agent_base import AgentBase
from models.schemas import AgentMessage

logger = logging.getLogger(__name__)


class TabularProcessorAgent(AgentBase):
    """Parses tabular files and returns DataFrames with schema information."""

    def __init__(self) -> None:
        super().__init__(
            agent_name="tabular_processor",
            description="Parses CSV, XLSX, XLS, Parquet, Feather, TSV, and ODS files into structured DataFrames.",
        )

    async def handle(self, message: AgentMessage) -> dict[str, Any]:
        """Process a tabular file and return schema + preview data."""
        payload = message.payload
        file_path: str = payload.get("file_path", "")
        mime_type: str = payload.get("mime_type", "")
        extension: str = payload.get("extension", "")
        content_info: dict[str, Any] = payload.get("content_info", {})

        if not file_path or not os.path.exists(file_path):
            return {"error": f"File not found: {file_path}"}

        ext = extension or os.path.splitext(file_path)[1].lower()

        try:
            if ext in (".csv", ".tsv"):
                return await self._parse_csv(file_path, ext, content_info)
            elif ext in (".xlsx", ".xls"):
                return await self._parse_excel(file_path, ext)
            elif ext == ".parquet":
                return await self._parse_parquet(file_path)
            elif ext == ".feather":
                return await self._parse_feather(file_path)
            elif ext == ".ods":
                return await self._parse_ods(file_path)
            else:
                # Attempt CSV as fallback
                return await self._parse_csv(file_path, ext, content_info)
        except Exception as exc:
            self.logger.error("Failed to parse tabular file %s: %s", file_path, exc)
            return {"error": str(exc), "file_path": file_path}

    async def _parse_csv(
        self, file_path: str, ext: str, content_info: dict[str, Any]
    ) -> dict[str, Any]:
        """Parse CSV/TSV using Polars with encoding and delimiter detection."""
        import polars as pl

        delimiter = content_info.get("detected_delimiter", "," if ext != ".tsv" else "\t")
        encoding = content_info.get("encoding", "utf-8")
        has_header = content_info.get("likely_header", True)

        try:
            df = pl.read_csv(
                file_path,
                separator=delimiter,
                encoding=encoding if encoding in ("utf-8", "utf-8-sig") else "utf8",
                has_header=has_header,
                infer_schema_length=10000,
                try_parse_dates=True,
                ignore_errors=True,
            )
        except Exception:
            # Fallback: try with latin-1 encoding
            self.logger.info("Retrying CSV parse with latin-1 encoding")
            df = pl.read_csv(
                file_path,
                separator=delimiter,
                encoding="utf8",
                has_header=has_header,
                ignore_errors=True,
            )

        schema = [
            {"name": col, "dtype": str(df[col].dtype)}
            for col in df.columns
        ]

        preview_rows = df.head(100).to_dicts()

        return {
            "format": "csv" if ext != ".tsv" else "tsv",
            "delimiter": delimiter,
            "encoding": encoding,
            "has_header": has_header,
            "row_count": df.height,
            "column_count": df.width,
            "schema": schema,
            "preview": preview_rows,
            "file_path": file_path,
        }

    async def _parse_excel(self, file_path: str, ext: str) -> dict[str, Any]:
        """Parse XLSX/XLS using openpyxl, handling multi-sheet workbooks."""
        import polars as pl

        sheets_data: list[dict[str, Any]] = []

        try:
            # Get sheet names
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception:
            sheet_names = ["Sheet1"]

        for sheet_name in sheet_names:
            try:
                df = pl.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    infer_schema_length=10000,
                )

                schema = [
                    {"name": col, "dtype": str(df[col].dtype)}
                    for col in df.columns
                ]

                sheets_data.append({
                    "sheet_name": sheet_name,
                    "row_count": df.height,
                    "column_count": df.width,
                    "schema": schema,
                    "preview": df.head(100).to_dicts(),
                })
            except Exception as exc:
                self.logger.warning(
                    "Failed to parse sheet '%s': %s", sheet_name, exc
                )
                sheets_data.append({
                    "sheet_name": sheet_name,
                    "error": str(exc),
                })

        total_rows = sum(s.get("row_count", 0) for s in sheets_data)

        return {
            "format": ext.lstrip("."),
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "sheets": sheets_data,
            "total_rows": total_rows,
            "file_path": file_path,
        }

    async def _parse_parquet(self, file_path: str) -> dict[str, Any]:
        """Parse Parquet files using PyArrow."""
        import pyarrow.parquet as pq
        import polars as pl

        # Get metadata from PyArrow
        pf = pq.ParquetFile(file_path)
        metadata = pf.metadata

        df = pl.read_parquet(file_path)

        schema = [
            {"name": col, "dtype": str(df[col].dtype)}
            for col in df.columns
        ]

        return {
            "format": "parquet",
            "row_count": df.height,
            "column_count": df.width,
            "row_groups": metadata.num_row_groups if metadata else None,
            "schema": schema,
            "preview": df.head(100).to_dicts(),
            "file_path": file_path,
        }

    async def _parse_feather(self, file_path: str) -> dict[str, Any]:
        """Parse Feather/Arrow IPC files."""
        import polars as pl

        df = pl.read_ipc(file_path)

        schema = [
            {"name": col, "dtype": str(df[col].dtype)}
            for col in df.columns
        ]

        return {
            "format": "feather",
            "row_count": df.height,
            "column_count": df.width,
            "schema": schema,
            "preview": df.head(100).to_dicts(),
            "file_path": file_path,
        }

    async def _parse_ods(self, file_path: str) -> dict[str, Any]:
        """Parse ODS files using Polars."""
        import polars as pl

        try:
            df = pl.read_ods(file_path, infer_schema_length=10000)

            schema = [
                {"name": col, "dtype": str(df[col].dtype)}
                for col in df.columns
            ]

            return {
                "format": "ods",
                "row_count": df.height,
                "column_count": df.width,
                "schema": schema,
                "preview": df.head(100).to_dicts(),
                "file_path": file_path,
            }
        except Exception as exc:
            return {"error": f"ODS parsing failed: {exc}", "file_path": file_path}
