"""Export Agent - exports data in various formats."""

from __future__ import annotations

import csv
import json
import logging
import os
import sqlite3
from typing import Any
from uuid import uuid4

from core.agent_base import AgentBase
from models.schemas import AgentMessage

logger = logging.getLogger(__name__)

DATABASE_PATH = os.getenv("DATABASE_PATH", "./data/databases")
EXPORT_PATH = os.getenv("EXPORT_PATH", "./data/exports")


class ExportAgent(AgentBase):
    """Exports data from user databases to CSV, JSON, XLSX, or Markdown."""

    def __init__(self) -> None:
        super().__init__(
            agent_name="export_agent",
            description="Exports data to CSV, JSON, XLSX, Markdown, and SQL dump formats.",
        )

    async def handle(self, message: AgentMessage) -> dict[str, Any]:
        payload = message.payload
        user_id = payload.get("user_id", "")
        table_name = payload.get("table_name", "")
        file_id = payload.get("file_id", "")
        export_format = payload.get("format", "csv")
        sql_query = payload.get("sql", "")
        # Multi-sheet: list of {name, table_name} dicts
        tables = payload.get("tables", [])
        # Per-sheet: export only this sheet
        sheet_name = payload.get("sheet_name", "")

        if not user_id:
            return {"error": "user_id required"}

        os.makedirs(os.path.join(EXPORT_PATH, user_id), exist_ok=True)

        # Multi-sheet export
        if tables and len(tables) > 1 and not sheet_name:
            return self._export_multi_sheet(user_id, tables, export_format, file_id)

        # Single sheet from multi-sheet file
        if sheet_name and tables:
            for t in tables:
                if t.get("name") == sheet_name:
                    table_name = t["table_name"]
                    break

        # Load data
        data = self._load_data(user_id, table_name, file_id, sql_query)
        if "error" in data:
            return data

        export_name = f"export_{table_name or file_id or 'data'}"
        return self._export_single(data, user_id, export_name, export_format)

    def _load_data(self, user_id: str, table_name: str, file_id: str, sql: str) -> dict[str, Any]:
        user_db = os.path.join(DATABASE_PATH, user_id, "user_data.db")
        if not os.path.exists(user_db):
            return {"error": "No user database found"}

        conn = sqlite3.connect(user_db)
        conn.row_factory = sqlite3.Row
        try:
            if sql:
                upper = sql.upper().strip()
                if not upper.startswith("SELECT"):
                    return {"error": "Only SELECT queries allowed for export"}
                cursor = conn.execute(sql)
            else:
                tbl = table_name or (f"file_{file_id.replace('-', '_')}" if file_id else "")
                if not tbl:
                    return {"error": "No table specified"}
                cursor = conn.execute(f'SELECT * FROM "{tbl}" LIMIT 50000')

            columns = [d[0] for d in cursor.description] if cursor.description else []
            rows = [dict(r) for r in cursor.fetchall()]
            return {"columns": columns, "rows": rows}
        except Exception as exc:
            return {"error": str(exc)}
        finally:
            conn.close()

    def _load_table(self, user_id: str, table_name: str) -> dict[str, Any]:
        """Load a single table's data."""
        return self._load_data(user_id, table_name, "", "")

    def _export_single(self, data: dict, user_id: str, name: str, fmt: str) -> dict[str, Any]:
        if fmt == "csv":
            return self._export_csv(data, user_id, name)
        elif fmt == "json":
            return self._export_json(data, user_id, name)
        elif fmt == "xlsx":
            return self._export_xlsx(data, user_id, name)
        elif fmt == "markdown":
            return self._export_markdown(data, user_id, name)
        return self._export_csv(data, user_id, name)

    def _export_multi_sheet(self, user_id: str, tables: list[dict], fmt: str, file_id: str) -> dict[str, Any]:
        """Export multiple sheets/tables into a single file."""
        export_name = f"export_{file_id or 'multi'}_{uuid4().hex[:8]}"
        sheet_data: dict[str, dict] = {}

        for t in tables:
            name = t.get("name", t.get("table_name", ""))
            tbl = t.get("table_name", "")
            data = self._load_table(user_id, tbl)
            if "error" not in data:
                sheet_data[name] = data

        if not sheet_data:
            return {"error": "No sheet data could be loaded"}

        if fmt == "xlsx":
            return self._export_xlsx_multi(sheet_data, user_id, export_name)
        elif fmt == "json":
            return self._export_json_multi(sheet_data, user_id, export_name)
        elif fmt == "csv":
            return self._export_csv_multi(sheet_data, user_id, export_name)

        return self._export_csv_multi(sheet_data, user_id, export_name)

    # --- Single-format exporters ---

    def _export_csv(self, data: dict, user_id: str, name: str) -> dict[str, Any]:
        path = os.path.join(EXPORT_PATH, user_id, f"{name}.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=data["columns"])
            writer.writeheader()
            writer.writerows(data["rows"])
        return {"path": path, "format": "csv", "row_count": len(data["rows"])}

    def _export_json(self, data: dict, user_id: str, name: str) -> dict[str, Any]:
        path = os.path.join(EXPORT_PATH, user_id, f"{name}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data["rows"], f, indent=2, default=str, ensure_ascii=False)
        return {"path": path, "format": "json", "row_count": len(data["rows"])}

    def _export_xlsx(self, data: dict, user_id: str, name: str) -> dict[str, Any]:
        """Export as real XLSX using openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            # Fallback to CSV if openpyxl not available
            logger.warning("openpyxl not available, falling back to CSV")
            return self._export_csv(data, user_id, name)

        path = os.path.join(EXPORT_PATH, user_id, f"{name}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        cols = data["columns"]
        bold = Font(bold=True)

        # Header
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.font = bold

        # Data rows
        for r_idx, row in enumerate(data["rows"], 2):
            for c_idx, col in enumerate(cols, 1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(col))

        # Auto-width (cap at 40 chars)
        for c_idx, col in enumerate(cols, 1):
            max_len = len(str(col))
            for row in data["rows"][:100]:
                val = row.get(col)
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = min(max_len + 2, 40)

        wb.save(path)
        return {"path": path, "format": "xlsx", "row_count": len(data["rows"])}

    def _export_markdown(self, data: dict, user_id: str, name: str) -> dict[str, Any]:
        path = os.path.join(EXPORT_PATH, user_id, f"{name}.md")
        cols = data["columns"]
        rows = data["rows"][:5000]
        with open(path, "w", encoding="utf-8") as f:
            f.write("| " + " | ".join(cols) + " |\n")
            f.write("| " + " | ".join(["---"] * len(cols)) + " |\n")
            for row in rows:
                f.write("| " + " | ".join(str(row.get(c, "")) for c in cols) + " |\n")
        return {"path": path, "format": "markdown", "row_count": len(rows)}

    # --- Multi-sheet exporters ---

    def _export_xlsx_multi(self, sheet_data: dict[str, dict], user_id: str, name: str) -> dict[str, Any]:
        """Export multiple sheets as real XLSX worksheets."""
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            return self._export_csv_multi(sheet_data, user_id, name)

        path = os.path.join(EXPORT_PATH, user_id, f"{name}.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        bold = Font(bold=True)
        total_rows = 0

        for sheet_name, data in sheet_data.items():
            # Sanitize sheet name (Excel max 31 chars, no special chars)
            safe_name = sheet_name[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
            ws = wb.create_sheet(title=safe_name or "Sheet")
            cols = data.get("columns", [])

            for c_idx, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=c_idx, value=col)
                cell.font = bold

            rows = data.get("rows", [])
            for r_idx, row in enumerate(rows, 2):
                for c_idx, col in enumerate(cols, 1):
                    ws.cell(row=r_idx, column=c_idx, value=row.get(col))

            total_rows += len(rows)

        wb.save(path)
        return {
            "path": path,
            "format": "xlsx",
            "row_count": total_rows,
            "sheets": list(sheet_data.keys()),
        }

    def _export_json_multi(self, sheet_data: dict[str, dict], user_id: str, name: str) -> dict[str, Any]:
        path = os.path.join(EXPORT_PATH, user_id, f"{name}.json")
        output = {sname: data.get("rows", []) for sname, data in sheet_data.items()}
        total_rows = sum(len(v) for v in output.values())
        with open(path, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, default=str, ensure_ascii=False)
        return {
            "path": path,
            "format": "json",
            "row_count": total_rows,
            "sheets": list(sheet_data.keys()),
        }

    def _export_csv_multi(self, sheet_data: dict[str, dict], user_id: str, name: str) -> dict[str, Any]:
        path = os.path.join(EXPORT_PATH, user_id, f"{name}.csv")
        total_rows = 0
        with open(path, "w", newline="", encoding="utf-8") as f:
            for sname, data in sheet_data.items():
                cols = data.get("columns", [])
                rows = data.get("rows", [])
                f.write(f"# --- Sheet: {sname} ---\n")
                writer = csv.DictWriter(f, fieldnames=cols)
                writer.writeheader()
                writer.writerows(rows)
                f.write("\n")
                total_rows += len(rows)

        return {
            "path": path,
            "format": "csv",
            "row_count": total_rows,
            "sheets": list(sheet_data.keys()),
        }
